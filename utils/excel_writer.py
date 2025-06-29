from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime

def criarArquivo (cotacoes, historico, selic, log):
    arquivo = Workbook()

    def formatarAba(aba, dados):
        if not dados:
            return None

        fonteNegrito = Font(bold=True)
        numeroEstilo = NamedStyle(name="numeroEstilo", number_format="0.00") #Variavel para setar pontos flutuantes em ate duas casas

        #Criando os cabeçalhos dos arquivos
        for indexColuna, coluna in enumerate(dados[0].keys(), start=1):
            cell = aba.cell(row = 1, column = indexColuna, value = coluna)
            cell.font = fonteNegrito
            cell.alignment = Alignment(horizontal = "center")

        #Criando as linhas de dados
        for indexLinha, itens in enumerate(dados, start=2):
            for indexColuna, valor in enumerate(itens.values(), start=1):
                cell = aba.cell(row = indexLinha, column = indexColuna, value = valor)
                if isinstance(valor, (int, float)):
                    cell.style = numeroEstilo

        #Ajustando a largura das colunas
        for indexColuna in range(1, len(dados[0])+1):
            letraColuna = get_column_letter(indexColuna)
            #aba.column_dimensions[letraColuna].auto_size = True
            aba.column_dimensions[letraColuna].width = 15

    ws_resumo = arquivo.active
    ws_resumo.title = "Resumo Atual"
    formatarAba(ws_resumo, cotacoes)

    # Aba 2: Histórico 30 dias do BTC
    ws_hist = arquivo.create_sheet("Histórico 30 dias")
    formatarAba(ws_hist, historico)

    # Aba 3: Data da pesquisa + porcetagem da Selic
    ws_selic = arquivo.create_sheet("Selic")
    formatarAba(ws_selic, selic)

    # Aba 4: Log de Execução
    ws_log = arquivo.create_sheet("Log de Execução")
    formatarAba(ws_log, log)

    # Salvar o arquivo
    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo.save(f"relatorio_{data_hora}.xlsx")
    print("[✓] Relatório salvo com sucesso!")